# Using openpyxl to work with Excel spreahsheets
# https://www.pyxll.com/blog/tools-for-working-with-excel-and-python/
# https://openpyxl.readthedocs.io/en/stable/tutorial.html

import openpyxl
# Open the spreadsheet
wb = openpyxl.load_workbook('PythonTest.xlsx')
# Print the sheet names in the spreadsheet
print(wb.sheetnames)
# Assign the 3rd sheet (by name) to an object
sheet =  wb['Sheet3']
print(sheet)
print(sheet.title)

# Create a new spreadsheet
# Assign the first sheet by name to an object
sheet = wb['Sheet1']
# Display the sheet information
print('Sheet A1 =  ', sheet['A1'])
# Display the value in cell A1
print('A1.value = ', sheet['A1'].value)
# Display the B column (column 2) values in a loop
print('\nB  Value')
for i in range(1, 5):
    print('B'+ str(i), sheet.cell(row=i, column=2).value)

# Change the value in B2
sheet['B2'] = 5

# Save the updated spreadsheet to a new spreadsheet
wb.save('example_copy.xlsx')
print('Save to example_copy.xlsx Done.')